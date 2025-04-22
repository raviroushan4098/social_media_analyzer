# reddit_profile_analyzer/views.py
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views import View
from .forms import UploadUsernamesCSVForm
import csv
import requests
from datetime import datetime
from nltk.sentiment import SentimentIntensityAnalyzer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string  # For rendering HTML snippets

class RedditProfileAnalyzerView(View):
    template_name = 'reddit_profile_analyzer/upload_form.html'

    def get(self, request):
        form = UploadUsernamesCSVForm()
        return render(request, self.template_name, {'form': form})

    @method_decorator(csrf_exempt)
    def post(self, request):
        form = UploadUsernamesCSVForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            usernames = [row[0].strip() for row in csv.reader(csv_file.read().decode('utf-8').splitlines()) if row]
            request.session['last_uploaded_usernames'] = usernames  # Store usernames in session
            analysis_results = []
            sia = SentimentIntensityAnalyzer()
            headers = {'User-Agent': 'Mozilla/5.0'}

            for username in usernames:
                url = f"https://www.reddit.com/user/{username}/about.json"
                try:
                    res = requests.get(url, headers=headers, timeout=10)
                    if res.status_code == 200:
                        data = res.json().get('data', {})
                        name = data.get('name', 'N/A')
                        link = f"https://www.reddit.com/user/{name}"
                        post_karma = data.get('link_karma', 0)
                        comment_karma = data.get('comment_karma', 0)
                        total_karma = post_karma + comment_karma
                        created = datetime.utcfromtimestamp(data.get('created_utc', 0)).strftime('%Y-%m-%d')
                        account_age_days = (datetime.utcnow() - datetime.utcfromtimestamp(data.get('created_utc', 0))).days
                        karma_per_day = total_karma

                        posts_url = f"https://www.reddit.com/user/{username}/submitted.json"
                        comments_url = f"https://www.reddit.com/user/{username}/comments.json"
                        posts = self.fetch_reddit_data(posts_url, headers)
                        comments = self.fetch_reddit_data(comments_url, headers)

                        items = self.process_items(posts, "P", sia, karma_per_day) + self.process_items(comments, "C", sia, karma_per_day)

                        analysis_results.append({
                            "name": name,
                            "link": link,
                            "created": created,
                            "karma": total_karma,
                            "karma_per_day": karma_per_day,
                            "account_age_days": account_age_days,
                            "items": items
                        })
                    else:
                        analysis_results.append({"error": f"Failed to retrieve data for {username}"})
                except Exception as e:
                    analysis_results.append({"error": f"Error processing {username}: {e}"})

            return render(request, 'reddit_profile_analyzer/analysis_results.html', {'results': analysis_results})
        else:
            return render(request, self.template_name, {'form': form, 'error': 'Invalid CSV file.'})

    def fetch_reddit_data(self, url, headers):
        try:
            res = requests.get(url, headers=headers, timeout=10)
            if res.status_code == 200:
                return res.json().get('data', {}).get('children', [])
        except Exception as e:
            print(f"Error fetching data: {e}")
        return []

    def process_items(self, items, item_type, sia, karma_per_day):
        processed_items = []
        total_comments = len(items) if item_type == "C" else 0

        for item in items:
            data = item.get('data', {})
            author = data.get('author', 'N/A')
            title_or_body = data.get('title', '') if item_type == "P" else data.get('body', '')
            sentiment = sia.polarity_scores(title_or_body)
            positive_count = 1 if sentiment['pos'] > 0.05 else 0
            negative_count = 1 if sentiment['neg'] < -0.05 else 0

            # Convert UTC timestamp to readable date
            created_utc = data.get('created_utc', 0)
            created_date = datetime.utcfromtimestamp(created_utc).strftime('%Y-%m-%d ') if created_utc else 'N/A'

            processed_items.append({
                "Sno": len(processed_items) + 1,
                "Type": item_type,
                "Name": data.get('title', 'N/A') if item_type == "P" else title_or_body[:100],
                "Account": author,
                "Account-Link": f"https://www.reddit.com/user/{author}",
                "Subreddit": data.get('subreddit', 'N/A'),
                "Subreddit(O/G)": "G",
                "Comment/Post-Link": f"https://www.reddit.com{data.get('permalink', '')}",
                "Related-To-LPU": "Yes" if "lpu" in title_or_body.lower() else "No",
                "Views": "N/A",
                "Post-Upvote-Count": data.get('ups', 0),
                "No-of-comments-in-the-post": data.get('num_comments', 0) if item_type == "P" else "N/A",
                "Positive-Comments-Count": positive_count,
                "Negative-Comments-Count": negative_count,
                "No-of-Comments-": total_comments if item_type == "C" else "N/A",
                "Comment-Upvote-Count": data.get('ups', 0) if item_type == "C" else "N/A",
                "Karma-Points-/-Day": karma_per_day,
                "Created-Date": created_date  # Add the date field
            })

        return processed_items

def export_to_excel_view(request):
    usernames = request.session.get('last_uploaded_usernames')
    if not usernames:
        return HttpResponse("No analysis data to export.", status=400)

    filename_from_prompt = request.GET.get('filename', 'Reddit_Analysis') + '.xlsx'

    all_items = []
    sia = SentimentIntensityAnalyzer()
    headers = {'User-Agent': 'Mozilla/5.0'}
    analyzer = RedditProfileAnalyzerView()  # Create instance to use its methods

    for username in usernames:
        url = f"https://www.reddit.com/user/{username}/about.json"
        try:
            res = requests.get(url, headers=headers, timeout=10)
            if res.status_code == 200:
                data = res.json().get('data', {})
                karma_per_day = (data.get('link_karma', 0) + data.get('comment_karma', 0))

                posts_url = f"https://www.reddit.com/user/{username}/submitted.json"
                comments_url = f"https://www.reddit.com/user/{username}/comments.json"
                
                # Use the analyzer instance to fetch data
                posts = analyzer.fetch_reddit_data(posts_url, headers)
                comments = analyzer.fetch_reddit_data(comments_url, headers)
                
                # Process items using the analyzer instance
                all_items.extend(analyzer.process_items(posts, "P", sia, karma_per_day))
                all_items.extend(analyzer.process_items(comments, "C", sia, karma_per_day))
        except Exception as e:
            print(f"Error during export data retrieval: {e}")

    # Create workbook only if we have items
    if not all_items:
        return HttpResponse("No data to export.", status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reddit Analysis Report"

    headers = [
        "Sno", "Type", "Title", "Account", "Account-Link", "Subreddit",
        "Subreddit(O/G)", "Comment/Post-Link", "Related-To-LPU", "Views",
        "Post-Upvote-Count", "No-of-comments-in-the-post",
        "Positive-Comments-Count", "Negative-Comments-Count", "No-of-Comments-",
        "Comment-Upvote-Count", "Karma-Points-/-Day", "Created-Date"  # Add Created-Date to headers
    ]
    
    # Add headers
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Add data rows
    for item in all_items:
        ws.append([
            item["Sno"], 
            item["Type"], 
            item["Name"],
            item["Account"], 
            item["Account-Link"],
            item["Subreddit"], 
            item["Subreddit(O/G)"], 
            item["Comment/Post-Link"],
            item["Related-To-LPU"], 
            item["Views"], 
            item["Post-Upvote-Count"],
            item["No-of-comments-in-the-post"], 
            item["Positive-Comments-Count"],
            item["Negative-Comments-Count"], 
            item["No-of-Comments-"],
            item["Comment-Upvote-Count"], 
            item["Karma-Points-/-Day"],
            item["Created-Date"]  # Add Created-Date to row data
        ])

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename_from_prompt}"'
    wb.save(response)
    return response

def analysis_results_view(request, results):
    return render(request, 'reddit_profile_analyzer/analysis_results.html', {'results': results})
def dashboard_view(request):
    return render(request, 'reddit_profile_analyzer/dashboard.html')